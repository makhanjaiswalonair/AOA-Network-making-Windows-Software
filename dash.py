from PyQt5 import QtCore, QtGui, QtWidgets
import openpyxl
from PyQt5.QtWidgets import QMessageBox
from spread import Ui_Spreadsheet
from Settings import SettingsDialog
from LoginWindow import LoginDialog
from CustomMessageBox import CustomMessageBox
import configparser,os,sys

class Ui_Dashboard(object):
    def resource_path(self,relative_path):
        """ Get absolute path to resource, works for dev and for PyInstaller """
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
        return os.path.join(base_path, relative_path)
    
    def import_file(self,name):
      import sys
      options = QtWidgets.QFileDialog.Options()
      file_path, _ = QtWidgets.QFileDialog.getOpenFileName(None, "Import Excel File", "", "Excel Files (*.xlsx)", options=options)
      if file_path:
        self.imported_file_path = file_path
        wb = openpyxl.load_workbook(file_path)
        sheet_names = wb.sheetnames

        # If there's more than one sheet, ask the user to select one
        if len(sheet_names) > 1:
            sheet_name, ok = QtWidgets.QInputDialog.getItem(None, "Select Worksheet", "Select the worksheet to import", sheet_names, 0, False)
            if not ok:
                return  # User cancelled the selection
        else:
            sheet_name = sheet_names[0]

        ws = wb[sheet_name]
        data = []

        for row in ws.iter_rows(values_only=True):
            row = list(row)
            if row[-1] is None:
                row[-1] = ""
            data.append(row)
        self.populate_spreadsheet(data, name)

    def populate_spreadsheet(self, data,name):
        self.window = QtWidgets.QMainWindow()
        self.ui = Ui_Spreadsheet()
        self.ui.setupUi(self.window)
        self.ui.label.setText(name)
        self.ui.populate_table(data)
        self.window.show()
        self.window.showMaximized()
        # QtWidgets.QApplication.instance().activeWindow().close()
        self.main_window.close()

    def create(self,name):
        self.window =QtWidgets.QMainWindow()
        self.ui = Ui_Spreadsheet()
        self.ui.setupUi(self.window)
        self.ui.label.setText(name)
        self.window.show()
        self.window.showMaximized()
        # QtWidgets.QApplication.instance().activeWindow().close()
        self.main_window.close()

    def create_file(self):
        # Show input dialog to get file name
        name, ok = QtWidgets.QInputDialog.getText(self.centralwidget, 'Create File', 'Enter name for your file:')
        
        if ok and name:
            self.create(name)
        elif ok:
            QtWidgets.QMessageBox.warning(self.centralwidget, 'No Name Entered', 'Please enter a valid name.')
        else:
            return

    def import_file_dialog(self):
        # Show input dialog to get file name
        name, ok = QtWidgets.QInputDialog.getText(self.centralwidget, 'Import File', 'Enter name for your file:')
        
        if ok and name:
            self.import_file(name)
        elif ok:
            QtWidgets.QMessageBox.warning(self.centralwidget, 'No Name Entered', 'Please enter a valid name.')
        else:
            return

    def open(self):
      import sys
      options = QtWidgets.QFileDialog.Options()
      file_path, _ = QtWidgets.QFileDialog.getOpenFileName(None, "Import Excel File", "", "Excel Files (*.xlsx)", options=options)
      if file_path:
        self.imported_file_path = file_path
        wb = openpyxl.load_workbook(file_path)
        sheet_names = wb.sheetnames

        # If there's more than one sheet, ask the user to select one
        if len(sheet_names) > 1:
            sheet_name, ok = QtWidgets.QInputDialog.getItem(None, "Select Worksheet", "Select the worksheet to import", sheet_names, 0, False)
            if not ok:
                return  # User cancelled the selection
        else:
            sheet_name = sheet_names[0]

        ws = wb[sheet_name]
        data = []

        for row in ws.iter_rows(values_only=True):
            row = list(row)
            if row[-1] is None:
                row[-1] = ""
            data.append(row)
        self.populate_spreadsheet(data, sheet_name)

    def open_settings(self):
        dialog = SettingsDialog()
        dialog.exec_()

    def handle_button(self):
        if self.LogOut.text() == "Login":
            self.login_window.exec_()
        elif self.LogOut.text() == "Logout":
            self.open_logout_confirmation()
        else:
            raise Exception("Invalid button text")
        
    def handle_login_successful(self, username, firstName):
        self.username = username
        self.firstName = firstName
        self.Myaccount.setText(f" {self.firstName}")
        self.LogOut.setText("Logout")

    def open_logout_confirmation(self):
        if CustomMessageBox.question(self.widget):
            # Reset user data
            self.username = None
            self.firstName = "Guest"
            self.Myaccount.setText(f" {self.firstName}")
            self.LogOut.setText("Login")

            # Remove user data from config.ini
            config = configparser.ConfigParser()
            config.read('config.ini')

            if 'User' in config:
                config.remove_option('User', 'username')
                config.remove_option('User', 'firstName')
                config.remove_option('User', 'token')

            with open('config.ini', 'w') as configfile:
                config.write(configfile)

    def load_user_data(self):
        if os.path.exists('config.ini'):
            config = configparser.ConfigParser()
            config.read('config.ini')

            if 'User' in config:
                username = config.get('User', 'username', fallback=None)
                firstName = config.get('User', 'firstName', fallback=None)
                token = config.get('User', 'token', fallback=None)

                # Check if any required field is missing
                if username is None or firstName is None or token is None:
                    return None

                return username, firstName, token

        return None

    
    def setupUi(self, Dashboard):
        self.main_window = Dashboard
        screen = QtWidgets.QDesktopWidget().screenGeometry()
        screen_width = screen.width()
        screen_height = screen.height()

        Dashboard.setObjectName("Dashboard")
        Dashboard.resize(screen_width, screen_height)
        Dashboard.setStyleSheet("background-color: rgb(235, 255, 255);")
        self.centralwidget = QtWidgets.QWidget(Dashboard)
        self.centralwidget.setObjectName("centralwidget")
        # self.centralwidget.setStyleSheet("background-color: blue;")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.centralwidget)
        self.centralwidget.setLayout(self.horizontalLayout)

        ## left menu
        self.left=QtWidgets.QWidget(self.centralwidget)
        self.horizontalLayout.addWidget(self.left)
        self.left.setStyleSheet("background-color: black;")
        self.leftverticalLayout = QtWidgets.QVBoxLayout(self.left)
        self.left.setLayout(self.leftverticalLayout)

        self.Home = QtWidgets.QPushButton(self.left)
        # self.Home.setGeometry(QtCore.QRect(40, 230, 201, 131))
        font = QtGui.QFont()
        font.setPointSize(15)
        self.Home.setFont(font)
        self.Home.setStyleSheet("background-color: black;\n"
"color: white;")
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap(self.resource_path("dashboard_images/home.ico")), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.Home.setIcon(icon1)
        self.Home.setIconSize(QtCore.QSize(45, 75))
        self.Home.setObjectName("Home")
        self.leftverticalLayout.addWidget(self.Home)

        self.Myaccount = QtWidgets.QPushButton(self.left)
        # self.Myaccount.setGeometry(QtCore.QRect(60, 80, 211, 101))
        font = QtGui.QFont()
        font.setPointSize(15)
        self.Myaccount.setFont(font)
        self.Myaccount.setStyleSheet("background-color: black;\n"
"color: white;")
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(self.resource_path("dashboard_images/user.ico")), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.Myaccount.setIcon(icon)
        self.Myaccount.setIconSize(QtCore.QSize(45, 75))
        self.Myaccount.setObjectName("Myaccount")
        self.leftverticalLayout.addWidget(self.Myaccount)

        self.Settings = QtWidgets.QPushButton(self.left)
        # self.Settings.setGeometry(QtCore.QRect(80, 440, 151, 101))
        font = QtGui.QFont()
        font.setPointSize(15)
        self.Settings.setFont(font)
        self.Settings.setStyleSheet("background-color: black;\n"
"color: white;")
        icon2 = QtGui.QIcon()
        icon2.addPixmap(QtGui.QPixmap(self.resource_path("dashboard_images/settings.ico")), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.Settings.setIcon(icon2)
        self.Settings.setIconSize(QtCore.QSize(45, 75))
        self.Settings.setObjectName("Settings")
        self.Settings.clicked.connect(self.open_settings)
        self.leftverticalLayout.addWidget(self.Settings)

        self.LogOut = QtWidgets.QPushButton(self.left)
        # self.LogOut.setGeometry(QtCore.QRect(70, 630, 151, 111))
        font = QtGui.QFont()
        font.setPointSize(15)
        self.LogOut.setFont(font)
        self.LogOut.setStyleSheet("background-color: black;\n"
"color: white;")
        icon3 = QtGui.QIcon()
        icon3.addPixmap(QtGui.QPixmap(self.resource_path("dashboard_images/logout.ico")), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.LogOut.setIcon(icon3)
        self.LogOut.setIconSize(QtCore.QSize(45, 75))
        self.LogOut.setObjectName("Logout")
        self.LogOut.clicked.connect(self.handle_button)        
        self.login_window = LoginDialog()
        self.login_window.login_successful.connect(self.handle_login_successful)
        self.leftverticalLayout.addWidget(self.LogOut)

        ## right menu
        self.right=QtWidgets.QWidget(self.centralwidget)
        self.horizontalLayout.addWidget(self.right)
        self.right.setStyleSheet("background-color: rgb(235, 255, 255);")
        self.rightverticalLayout = QtWidgets.QVBoxLayout(self.right)
        self.right.setLayout(self.rightverticalLayout)

        ## right1
        self.right1=QtWidgets.QWidget(self.right)
        self.rightverticalLayout.addWidget(self.right1)
        self.right1horizontalLayout = QtWidgets.QHBoxLayout(self.right1)
        self.right1.setLayout(self.right1horizontalLayout)

        self.Search = QtWidgets.QPushButton(self.right1)
        # self.Search.setGeometry(QtCore.QRect(530, 70, 201, 41))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.Search.setFont(font)
        self.Search.setStyleSheet("background-color: rgb(255, 255, 255);")
        icon4 = QtGui.QIcon()
        icon4.addPixmap(QtGui.QPixmap(self.resource_path("dashboard_images/search.ico")), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.Search.setIcon(icon4)
        self.Search.setIconSize(QtCore.QSize(45, 20))
        self.Search.setObjectName("Search")
        self.right1horizontalLayout.addWidget(self.Search)

        self.Line_Searchhere = QtWidgets.QLineEdit(self.right1)
        # self.Line_Searchhere.setGeometry(QtCore.QRect(810, 70, 341, 41))
        self.Line_Searchhere.setStyleSheet("background-color: white;\n"
"border-radius: 1 rem;")
        self.Line_Searchhere.setObjectName("Line_Searchhere")
        self.right1horizontalLayout.addWidget(self.Line_Searchhere)

        self.label_Searchhere = QtWidgets.QLabel(self.centralwidget)
        self.label_Searchhere.setStyleSheet("background-color: rgb(235, 255, 255);")
        self.right1horizontalLayout.addWidget(self.label_Searchhere)

        ## right2
        self.right2=QtWidgets.QWidget(self.right)
        self.rightverticalLayout.addWidget(self.right2)
        self.right2horizontalLayout = QtWidgets.QHBoxLayout(self.right2)
        self.right2.setLayout(self.right2horizontalLayout)
        self.Dash = QtWidgets.QLabel(self.right2)
        # self.Dash.setGeometry(QtCore.QRect(530, 160, 511, 61))
        font = QtGui.QFont()
        font.setPointSize(25)
        self.Dash.setFont(font)
        self.Dash.setObjectName("Dash")
        self.right2horizontalLayout.addWidget(self.Dash)

        ## right3
        self.right3=QtWidgets.QWidget(self.right)
        self.rightverticalLayout.addWidget(self.right3)
        self.right3horizontalLayout = QtWidgets.QHBoxLayout(self.right3)
        self.right3.setLayout(self.right3horizontalLayout)

        self.FileName1 = QtWidgets.QPushButton(self.right3)
        # self.FileName1.setGeometry(QtCore.QRect(120, 30, 191, 151))
        font = QtGui.QFont()
        font.setPointSize(15)
        self.FileName1.setFont(font)
        self.FileName1.setMouseTracking(False)
        self.FileName1.setStyleSheet("background-color: rgb(235, 255, 255);")
        icon5 = QtGui.QIcon()
        icon5.addPixmap(QtGui.QPixmap(self.resource_path("dashboard_images/file.ico")), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.FileName1.setIcon(icon5)
        self.FileName1.setIconSize(QtCore.QSize(70, 150))
        self.FileName1.setObjectName("pushButton_6")
        self.right3horizontalLayout.addWidget(self.FileName1)

        self.FileName2 = QtWidgets.QPushButton(self.right3)
        # self.FileName2.setGeometry(QtCore.QRect(380, 30, 201, 151))
        font = QtGui.QFont()
        font.setPointSize(15)
        self.FileName2.setFont(font)
        self.FileName2.setMouseTracking(False)
        self.FileName2.setIcon(icon5)
        self.FileName2.setIconSize(QtCore.QSize(70, 150))
        self.FileName2.setObjectName("pushButton_7")
        self.right3horizontalLayout.addWidget(self.FileName2)

        self.FileName3 = QtWidgets.QPushButton(self.right3)
        # self.pushButton_8.setGeometry(QtCore.QRect(650, 30, 191, 151))
        font = QtGui.QFont()
        font.setPointSize(15)
        self.FileName3.setFont(font)
        self.FileName3.setMouseTracking(False)
        self.FileName3.setIcon(icon5)
        self.FileName3.setIconSize(QtCore.QSize(70, 150))
        self.FileName3.setObjectName("pushButton_8")
        self.right3horizontalLayout.addWidget(self.FileName3)

        self.FileName4 = QtWidgets.QPushButton(self.right3)
        # self.pushButton_13.setGeometry(QtCore.QRect(920, 30, 211, 151))
        font = QtGui.QFont()
        font.setPointSize(15)
        self.FileName4.setFont(font)
        self.FileName4.setMouseTracking(False)
        self.FileName4.setIcon(icon5)
        self.FileName4.setIconSize(QtCore.QSize(70, 150))
        self.FileName4.setObjectName("pushButton_13")
        self.right3horizontalLayout.addWidget(self.FileName4)

        self.More = QtWidgets.QPushButton(self.right3)
        # self.pushButton_9.setGeometry(QtCore.QRect(1210, 30, 101, 151))
        self.More.setText("")
        icon6 = QtGui.QIcon()
        icon6.addPixmap(QtGui.QPixmap(self.resource_path("dashboard_images/menu.ico")), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.More.setIcon(icon6)
        self.More.setIconSize(QtCore.QSize(70, 150))
        self.More.setObjectName("pushButton_9")
        self.right3horizontalLayout.addWidget(self.More)

        ## right4
        self.right4=QtWidgets.QWidget(self.right)
        self.rightverticalLayout.addWidget(self.right4)
        self.right4horizontalLayout = QtWidgets.QHBoxLayout(self.right4)
        self.right4.setLayout(self.right4horizontalLayout)

        self.Create = QtWidgets.QPushButton(self.right4)
        # self.Create.setGeometry(QtCore.QRect(100, 30, 211, 181))
        font = QtGui.QFont()
        font.setPointSize(15)
        self.Create.setFont(font)
        icon7 = QtGui.QIcon()
        icon7.addPixmap(QtGui.QPixmap(self.resource_path("dashboard_images/create.ico")), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.Create.setIcon(icon7)
        self.Create.setIconSize(QtCore.QSize(100, 150))
        self.Create.setObjectName("Create")
        self.right4horizontalLayout.addWidget(self.Create)
        self.Create.clicked.connect(self.create_file)

        self.Import = QtWidgets.QPushButton(self.right4)
        # self.Import.setGeometry(QtCore.QRect(510, 30, 201, 181))
        font = QtGui.QFont()
        font.setPointSize(15)
        self.Import.setFont(font)
        icon8 = QtGui.QIcon()
        icon8.addPixmap(QtGui.QPixmap(self.resource_path("dashboard_images/download.ico")), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.Import.setIcon(icon8)
        self.Import.setIconSize(QtCore.QSize(100, 150))
        self.Import.setObjectName("Import")
        self.right4horizontalLayout.addWidget(self.Import)
        self.Import.clicked.connect(self.import_file_dialog)

        self.Open = QtWidgets.QPushButton(self.right4)
        # self.Open.setGeometry(QtCore.QRect(920, 30, 211, 181))
        font = QtGui.QFont()
        font.setPointSize(15)
        self.Open.setFont(font)
        icon9 = QtGui.QIcon()
        icon9.addPixmap(QtGui.QPixmap(self.resource_path("dashboard_images/open.ico")), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.Open.setIcon(icon9)
        self.Open.setIconSize(QtCore.QSize(100, 150))
        self.Open.setObjectName("Open")
        self.right4horizontalLayout.addWidget(self.Open)
        self.Open.clicked.connect(self.open)

        self.rightverticalLayout.setStretchFactor(self.right1, 1)
        self.rightverticalLayout.setStretchFactor(self.right2, 2)
        self.rightverticalLayout.setStretchFactor(self.right3, 6)
        self.rightverticalLayout.setStretchFactor(self.right4, 6)

        ## end
        self.end=QtWidgets.QWidget(self.centralwidget)
        self.end.setStyleSheet("background-color: rgb(235, 255, 255);")
        self.horizontalLayout.addWidget(self.end)

        self.horizontalLayout.setStretchFactor(self.left, 1)
        self.horizontalLayout.setStretchFactor(self.right, 5)
        self.horizontalLayout.setStretchFactor(self.end, 1)
        
        Dashboard.setCentralWidget(self.centralwidget)

        self.retranslateUi(Dashboard)
        QtCore.QMetaObject.connectSlotsByName(Dashboard)

    def retranslateUi(self, Dashboard):
        _translate = QtCore.QCoreApplication.translate
        Dashboard.setWindowTitle(_translate("Dashboard", "Dashboard"))
        self.Myaccount.setText(_translate("Dashboard", " My Account"))
        self.Home.setText(_translate("Dashboard", "Dashboard"))
        self.Settings.setText(_translate("Dashboard", "Settings"))
        self.LogOut.setText(_translate("Dashboard", "Logout"))
        self.Search.setText(_translate("Dashboard", "Search"))
        self.Dash.setText(_translate("Dashboard", "Dashboard"))
        self.FileName1.setText(_translate("Dashboard", "File Name"))
        self.FileName2.setText(_translate("Dashboard", "File Name"))
        self.FileName3.setText(_translate("Dashboard", "File Name"))
        self.FileName4.setText(_translate("Dashboard", "File Name"))
        self.Create.setText(_translate("Dashboard", "Create"))
        self.Import.setText(_translate("Dashboard", "Import"))
        self.Open.setText(_translate("Dashboard", "Open"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Dashboard = QtWidgets.QMainWindow()
    Dashboard.showMaximized()
    ui = Ui_Dashboard()
    ui.setupUi(Dashboard)
    user_data = ui.load_user_data()
    if user_data:
        username, firstName, token = user_data
        ui.handle_login_successful(username, firstName)
    Dashboard.show()
    sys.exit(app.exec_())